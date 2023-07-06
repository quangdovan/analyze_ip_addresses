import requests
import check_return_api
import pandas as pd
from tkinter.messagebox import showinfo
from openpyxl import load_workbook
from openpyxl import Workbook

def analyze_file_hash(input_file, output_file):
    # Mở file Excel hiện có hoặc tạo mới nếu chưa tồn tại
    try:
        workbook = load_workbook(output_file)
    except FileNotFoundError:
        workbook = Workbook()
        # Chọn sheet để ghi dữ liệu
    sheet = workbook.active
    sheet['A1'] = 'hash'
    sheet['B1'] = 'type_description'
    sheet['C1'] = 'bytehero_info'
    sheet['D1'] = 'tlsh'
    sheet['E1'] = 'vhash'
    sheet['F1'] = 'names'
    sheet['G1'] = 'signature_info'
    sheet['H1'] = 'suggested_threat_label'
    sheet['H1'] = 'last_analysis_stats'
    sheet['J1'] = 'full'

    # Lưu lại file Excel
    workbook.save(output_file)
    NA = "N/A"
    # Đọc tệp Excel và tạo DataFrame
    df = pd.read_excel(input_file)

    first_column_name = df.columns[0]
    if first_column_name != 'hash':
        showinfo("Thông báo", "Kiểm tra lại tên cột của file đầu vào!")

    # Duyệt qua các IP address trong DataFrame
    for hash in df['hash']:
        # Gửi yêu cầu đến VirusTotal API
        url = 'https://www.virustotal.com/api/v3/files/%s' % hash

        headers = {'x-apikey': '66a11c172d28c9230514ef565f2ec627bddbdf77d9dce3c9b7bd95b00ed6d491'}
        response = requests.request("GET", url, headers=headers)
        try:
            data = response.json()
        except:
            showinfo("Thông báo", "Kiểm tra lại kết nối!")
        check = check_return_api.check_return_api(data)

        # Kiểm tra mã phản hồi
        if response.status_code == 200 and check == "ok":
            try:
                type_description = data['data']['attributes']['type_description']
            except:
                type_description = NA
            try:
                bytehero_info = data['data']['attributes']['bytehero_info']
            except: 
                bytehero_info = NA
            try:
                tlsh = data['data']['attributes']['tlsh']
            except:
                tlsh = NA   
            try:
                vhash = data['data']['attributes']['vhash']
            except:
                vhash = NA
            try:
                names = data['data']['attributes']['names']
            except:
                names = NA
            try:
                signature_info = data['data']['attributes']['signature_info']
            except:
                signature_info = NA
            try:
                suggested_threat_label = data['data']['attributes']['popular_threat_classification']['suggested_threat_label']
            except: 
                suggested_threat_label = NA
            try:
                last_analysis_stats = data['data']['attributes']['last_analysis_stats']
            except:
                last_analysis_stats = NA
            
        else:
            type_description = NA
            bytehero_info = NA
            tlsh = NA
            vhash = NA
            names = NA
            signature_info = NA
            suggested_threat_label = NA
            last_analysis_stats = NA
            # Thêm kết quả phân tích vào danh sách

        next_row = sheet.max_row + 1

        sheet.cell(row=next_row, column=1).value = str(hash)
        sheet.cell(row=next_row, column=2).value = str(type_description)  # Sử dụng chỉ số cột thay vì tên cột
        sheet.cell(row=next_row, column=3).value = str(bytehero_info)
        sheet.cell(row=next_row, column=4).value = str(tlsh)
        sheet.cell(row=next_row, column=5).value = str(vhash)
        sheet.cell(row=next_row, column=6).value = str(names)
        sheet.cell(row=next_row, column=7).value = str(signature_info)
        sheet.cell(row=next_row, column=8).value = str(suggested_threat_label)
        sheet.cell(row=next_row, column=9).value = str(last_analysis_stats)
        try:
            sheet.cell(row=next_row, column=10).value = str(data)
        except:
            sheet.cell(row=next_row, column=10).value = str("Lỗi phản hồi từ máy chủ")

    # Lưu DataFrame vào tệp Excel đầu ra
    try:
        workbook.save(output_file)
        showinfo("Thông báo", "Phân tích hoàn thành!")
    except:
        showinfo("Thông báo", "Đã có lỗi xảy ra!")