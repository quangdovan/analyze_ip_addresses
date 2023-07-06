import requests
import check_return_api
import pandas as pd
from tkinter.messagebox import showinfo
from openpyxl import load_workbook
from openpyxl import Workbook


def analyze_ip_addresses(input_file, output_file):
    # Mở file Excel hiện có hoặc tạo mới nếu chưa tồn tại
    try:
        workbook = load_workbook(output_file)
    except FileNotFoundError:
        workbook = Workbook()
        # Chọn sheet để ghi dữ liệu
    sheet = workbook.active
    sheet['A1'] = 'ip_address'
    sheet['B1'] = 'regional_internet_registry'
    sheet['C1'] = 'network'
    sheet['D1'] = 'as_owner'
    sheet['E1'] = 'country'
    sheet['F1'] = 'last_analysis_stats'
    sheet['G1'] = 'full'

    # Lưu lại file Excel
    workbook.save(output_file)
    NA = "N/A"
    # Đọc tệp Excel và tạo DataFrame
    df = pd.read_excel(input_file)

    first_column_name = df.columns[0]
    if first_column_name != 'ip':
        showinfo("Thông báo", "Kiểm tra lại tên cột của file đầu vào!")

    # Duyệt qua các IP address trong DataFrame
    for ip_addresses in df['ip']:
        # Gửi yêu cầu đến VirusTotal API
        url = 'https://www.virustotal.com/api/v3/ip_addresses/%s' % ip_addresses

        headers = {'x-apikey': '66a11c172d28c9230514ef565f2ec627bddbdf77d9dce3c9b7bd95b00ed6d491'}
        response = requests.request("GET", url, headers=headers)
        try:
            data = response.json()
        except:
            showinfo("Thông báo", "Kiểm tra lại kết nối!")
        check = check_return_api.check_return_api(data)

        # Kiểm tra mã phản hồi
        if response.status_code == 200 and check == "ok":
            try:
                regional_internet_registry = data['data']['attributes']['regional_internet_registry']
            except:
                regional_internet_registry = NA
            try:
                network = data['data']['attributes']['network']
            except:
                network = NA
            try:
                as_owner = data['data']['attributes']['as_owner']
            except:
                as_owner = NA
            try:
                country = data['data']['attributes']['country']
            except:
                country = NA
            try:
                last_analysis_stats = data['data']['attributes']['last_analysis_stats']
            except:
                last_analysis_stats = NA
        else:
            regional_internet_registry = NA
            network = NA
            as_owner = data
            country = NA
            last_analysis_stats = NA

        # Xác định vị trí dòng tiếp theo để ghi dữ liệu
        
        next_row = sheet.max_row + 1

        sheet.cell(row=next_row, column=1).value = str(ip_addresses)
        sheet.cell(row=next_row, column=2).value = str(regional_internet_registry)  # Sử dụng chỉ số cột thay vì tên cột
        sheet.cell(row=next_row, column=3).value = str(network)
        sheet.cell(row=next_row, column=4).value = str(as_owner)
        sheet.cell(row=next_row, column=5).value = str(country)
        sheet.cell(row=next_row, column=6).value = str(last_analysis_stats)
        try:
            sheet.cell(row=next_row, column=7).value = str(data)
        except:
            sheet.cell(row=next_row, column=7).value = str("Lỗi phản hồi từ máy chủ")
    # Lưu DataFrame vào tệp Excel đầu ra
    try:
        workbook.save(output_file)
        showinfo("Thông báo", "Phân tích hoàn thành!")
    except:
        showinfo("Thông báo", "Đã có lỗi xảy ra!")