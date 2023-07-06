import pandas as pd
import requests
import check_return_api
from tkinter.messagebox import showinfo
from openpyxl import load_workbook
from openpyxl import Workbook

def analyze_domain(input_file, output_file):
    # Mở file Excel hiện có hoặc tạo mới nếu chưa tồn tại
    try:
        workbook = load_workbook(output_file)
    except FileNotFoundError:
        workbook = Workbook()
        # Chọn sheet để ghi dữ liệu
    sheet = workbook.active
    sheet = workbook.active
    sheet['A1'] = 'domain'
    sheet['B1'] = 'last_dns_records'
    sheet['C1'] = 'last_analysis_stats'
    sheet['D1'] = 'full'
    # Lưu lại file Excel
    workbook.save(output_file)
    NA = "N/A"
    # Đọc tệp Excel và tạo DataFrame
    df = pd.read_excel(input_file)

    first_column_name = df.columns[0]
    if first_column_name != 'domain':
        showinfo("Thông báo", "Kiểm tra lại tên cột của file đầu vào!")

    
    # Duyệt qua các IP address trong DataFrame
    for domain in df['domain']:
        # Gửi yêu cầu đến VirusTotal API
        url = 'https://www.virustotal.com/api/v3/domains/%s' % domain

        headers = {'x-apikey': '66a11c172d28c9230514ef565f2ec627bddbdf77d9dce3c9b7bd95b00ed6d491'}
        response = requests.request("GET", url, headers=headers)
        try:
            data = response.json()
        except:
            showinfo("Thông báo", "Kiểm tra lại kết nối!")
        check = check_return_api.check_return_api(data)

        # Kiểm tra mã phản hồi
        if response.status_code == 200 and check == "ok":
            try:
                last_dns_records = data['data']['attributes']['last_dns_records']
            except:
                last_dns_records = NA
            try:
                last_analysis_stats = data['data']['attributes']['last_analysis_stats']
            except:
                last_analysis_stats = NA
            # Lấy thông tin kết quả từ phản hồi

        else:
            last_dns_records = NA 
            last_analysis_stats = NA

    # Gán danh sách kết quả phân tích vào cột "Country" của DataFrame
        next_row = sheet.max_row + 1

        sheet.cell(row=next_row, column=1).value = str(domain)
        sheet.cell(row=next_row, column=2).value = str(last_dns_records)  # Sử dụng chỉ số cột thay vì tên cột
        sheet.cell(row=next_row, column=3).value = str(last_analysis_stats)
        sheet.cell(row=next_row, column=4).value = str(data)

    # Lưu DataFrame vào tệp Excel đầu ra
    try:
        workbook.save(output_file)
        showinfo("Thông báo", "Phân tích hoàn thành!")
    except:
        showinfo("Thông báo", "Đã có lỗi xảy ra!")